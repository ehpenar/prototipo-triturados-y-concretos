from __future__ import annotations

import re
import shutil
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, time, timedelta
from pathlib import Path
from typing import Any

import openpyxl


BASE_DIR = Path(__file__).resolve().parent

ORDENES_FILE = BASE_DIR / "Copia de ORDENES DE TRABAJO TYC.xlsx"
MATRIZ_FILE = BASE_DIR / "Copia de  Matriz de Seguimiento (respuestas).xlsx"
ACTIVIDADES_FILE = BASE_DIR / "Copia de REPORTE DE ACTIVIDADES MANTENIMIENTO (respuestas).xlsx"
TARGET_FILE = BASE_DIR / "HOJA RESUMEN FINANCIERO OTS.xlsx"

ORDENES_SHEET = "Respuestas de formulario 1"
MATRIZ_SHEET = "Respuestas de formulario 1"
ACTIVIDADES_SHEET = "Respuestas de formulario 1"
PERSONAL_SHEET = "PERSONAL"
FINANCIERO_SHEET = "Hoja 1"
DESTINO_SHEET = "Hoja 2"

OUTPUT_DIR = BASE_DIR / "outputs"


DEST_HEADERS = [
    "OT",
    "ESTATUS DE LA OT",
    "TIEMPO DE EJECUCION",
    "MANO OBRA",
    "#SP",
    "FECHA RECEPCION SP",
    "TIEMPO DE COMPRA",
    "ORDEN DE COMPRA",
    "TIEMPO APROBACION",
    "Estado Actual de la SP*",
    "VALOR DE LA COMPRA DE LA SP",
    "METODO_CRUCE_SP",
    "DETALLE_CRUCE_SP",
    "METODO_MANO_OBRA",
    "DETALLE_MANO_OBRA",
    "METODO_VALOR_COMPRA",
    "DETALLE_VALOR_COMPRA",
]


@dataclass
class MatchResult:
    sp: int | None
    method: str
    detail: str
    matriz_rows: list[dict[str, Any]]


def clean_header(value: Any) -> str:
    return "" if value is None else str(value).replace("\n", " ").strip()


def normalize_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if value == 0:
            return None
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    numbers = re.findall(r"\d+", text)
    if not numbers:
        return None
    number = int(numbers[0])
    return number or None


def normalize_ot(value: Any, row_number: int | None = None) -> int | None:
    if isinstance(value, str) and value.startswith('=CONCATENATE("OT-",ROW(A') and row_number:
        return row_number - 1
    return normalize_int(value)


def normalize_sp(value: Any) -> int | None:
    return normalize_int(value)


def parse_sp_mentions(*values: Any) -> list[int]:
    text = " ".join("" if v is None else str(v) for v in values)
    matches = re.findall(r"\bSP\s*#?\s*0*(\d{1,6})\b", text, flags=re.IGNORECASE)
    return [int(match) for match in matches if int(match) != 0]


def parse_duration_hours(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, timedelta):
        return value.total_seconds() / 3600
    if isinstance(value, time):
        return value.hour + value.minute / 60 + value.second / 3600
    if isinstance(value, (int, float)):
        # Excel stores fractions of a day for time-like values.
        return float(value) * 24 if 0 <= float(value) <= 1 else float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return 0.0
    if ":" in text:
        parts = [float(p) for p in text.split(":") if p != ""]
        if len(parts) >= 2:
            return parts[0] + parts[1] / 60 + (parts[2] / 3600 if len(parts) > 2 else 0)
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_money(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace("$", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        # In these files commas are often thousands separators.
        text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def days_between(start: Any, end: Any) -> float | None:
    if not isinstance(start, datetime) or not isinstance(end, datetime):
        return None
    return round((end - start).total_seconds() / 86400, 2)


def format_excel_value(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value)


def build_ot_status_value(orden_row: dict[str, Any]) -> str:
    marca_temporal = format_excel_value(orden_row.get("Marca temporal"))
    fecha_real_entrega = format_excel_value(orden_row.get("FECHA REAL ENTREGA"))
    return f"Marca temporal: {marca_temporal} | FECHA REAL ENTREGA: {fecha_real_entrega}"


def first_value(rows: list[dict[str, Any]], header: str) -> Any:
    for row in rows:
        value = row.get(header)
        if value not in (None, ""):
            return value
    return None


def rows_from_sheet(file_path: Path, sheet_name: str, data_only: bool = True) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(file_path, data_only=data_only, read_only=False)
    ws = wb[sheet_name]
    headers = [clean_header(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_number in range(2, ws.max_row + 1):
        row = {
            headers[col - 1]: ws.cell(row_number, col).value
            for col in range(1, ws.max_column + 1)
            if headers[col - 1]
        }
        if any(value not in (None, "") for value in row.values()):
            row["_row_number"] = row_number
            rows.append(row)
    return rows


def build_indexes() -> dict[str, Any]:
    ordenes = rows_from_sheet(ORDENES_FILE, ORDENES_SHEET, data_only=True)
    matriz = rows_from_sheet(MATRIZ_FILE, MATRIZ_SHEET, data_only=True)
    actividades = rows_from_sheet(ACTIVIDADES_FILE, ACTIVIDADES_SHEET, data_only=True)
    personal = rows_from_sheet(ACTIVIDADES_FILE, PERSONAL_SHEET, data_only=True)
    financiero = rows_from_sheet(TARGET_FILE, FINANCIERO_SHEET, data_only=True)

    ordenes_by_ot: dict[int, dict[str, Any]] = {}
    for row in ordenes:
        ot = normalize_ot(row.get("OT"), row.get("_row_number")) or normalize_ot(row.get("5"), row.get("_row_number"))
        if ot:
            row["_ot_norm"] = ot
            ordenes_by_ot[ot] = row

    matriz_by_ot: dict[int, list[dict[str, Any]]] = defaultdict(list)
    matriz_by_sp: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in matriz:
        ot = normalize_ot(row.get("OT"))
        sp = normalize_sp(row.get("NUMERO DE LA SP (solo el numero sin letras)*"))
        row["_ot_norm"] = ot
        row["_sp_norm"] = sp
        if ot:
            matriz_by_ot[ot].append(row)
        if sp:
            matriz_by_sp[sp].append(row)

    personal_rate_by_name = {
        str(row.get("i") or "").strip().upper(): parse_money(row.get("VALOR HORA"))
        for row in personal
        if row.get("i")
    }

    actividades_by_ot: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in actividades:
        ot = normalize_ot(row.get("OT")) or normalize_ot(row.get("ORDEN DE TRABAJO/REPORTE DE CAMPO"))
        row["_ot_norm"] = ot
        worker = str(row.get("BELMER AVALO ALZATE") or "").strip().upper()
        hours = parse_duration_hours(row.get("TIEMPO DE LA ACTIVIDAD"))
        rate = personal_rate_by_name.get(worker, 0.0)
        row["_worker"] = worker
        row["_hours"] = hours
        row["_rate"] = rate
        row["_labor_cost"] = hours * rate
        if ot:
            actividades_by_ot[ot].append(row)

    financiero_by_ot: dict[int, list[dict[str, Any]]] = defaultdict(list)
    financiero_by_sp: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in financiero:
        ot = normalize_ot(row.get("Número Orden de Trabajo"))
        sp = normalize_sp(row.get("SP"))
        row["_ot_norm"] = ot
        row["_sp_norm"] = sp
        if ot:
            financiero_by_ot[ot].append(row)
        if sp:
            financiero_by_sp[sp].append(row)

    return {
        "ordenes_by_ot": ordenes_by_ot,
        "matriz_by_ot": matriz_by_ot,
        "matriz_by_sp": matriz_by_sp,
        "actividades_by_ot": actividades_by_ot,
        "financiero_by_ot": financiero_by_ot,
        "financiero_by_sp": financiero_by_sp,
    }


def match_sp(ot: int, orden_row: dict[str, Any], indexes: dict[str, Any]) -> MatchResult:
    matriz_by_ot = indexes["matriz_by_ot"]
    matriz_by_sp = indexes["matriz_by_sp"]

    sp_from_column = normalize_sp(orden_row.get("sp"))
    if sp_from_column and sp_from_column in matriz_by_sp:
        return MatchResult(sp_from_column, "directo_columna_sp", "OT.sp cruza con Matriz.NUMERO DE LA SP", matriz_by_sp[sp_from_column])

    sp_mentions = parse_sp_mentions(
        orden_row.get("COMENTARIOS"),
        orden_row.get("DESCRIPCIÓN GENERAL DEL FALLO O DE LA SOLICTUD"),
        orden_row.get("Si no tiene el formato por favor describa claramente su solicitud (cantidad, dimensiones, material, etc)"),
    )
    for sp in sp_mentions:
        if sp in matriz_by_sp:
            return MatchResult(sp, "truco_sp_en_texto", f"SP {sp} extraida de comentarios/descripcion de la OT", matriz_by_sp[sp])

    if ot in matriz_by_ot:
        rows = matriz_by_ot[ot]
        sps = sorted({row["_sp_norm"] for row in rows if row.get("_sp_norm")})
        detail = f"Matriz.OT={ot}; SPs encontrados={sps}"
        return MatchResult(sps[0] if sps else None, "directo_matriz_ot", detail, rows)

    if sp_mentions:
        return MatchResult(sp_mentions[0], "truco_sp_en_texto_sin_matriz", f"SP {sp_mentions[0]} extraida, pero no existe en Matriz", [])

    if sp_from_column:
        return MatchResult(sp_from_column, "sp_columna_sin_matriz", f"OT.sp={sp_from_column}, pero no existe en Matriz", [])

    return MatchResult(None, "sin_cruce_sp", "No hay OT en Matriz ni SP detectable en la OT", [])


def labor_for_ot(ot: int, indexes: dict[str, Any]) -> tuple[float | None, str, str]:
    rows = indexes["actividades_by_ot"].get(ot, [])
    if not rows:
        return None, "sin_actividades", "No hay actividades con esta OT"
    total = sum(row["_labor_cost"] for row in rows)
    missing_rates = sum(1 for row in rows if row["_hours"] and not row["_rate"])
    detail = f"{len(rows)} actividades; {round(sum(row['_hours'] for row in rows), 2)} horas"
    if missing_rates:
        detail += f"; {missing_rates} actividades sin tarifa de PERSONAL"
    return round(total, 2), "directo_actividades_ot", detail


def purchase_value_for_ot(ot: int, sp: int | None, indexes: dict[str, Any]) -> tuple[float | None, str, str]:
    rows = indexes["financiero_by_ot"].get(ot, [])
    if not rows and sp:
        rows = indexes["financiero_by_sp"].get(sp, [])
        if rows:
            method = "directo_financiero_sp"
            detail_key = f"SP={sp}"
        else:
            return None, "sin_valor_compra", "No hay fila en Hoja 1 por OT ni SP"
    elif rows:
        method = "directo_financiero_ot"
        detail_key = f"OT={ot}"
    else:
        return None, "sin_valor_compra", "No hay fila en Hoja 1 por OT ni SP"

    total = 0.0
    for row in rows:
        total += parse_money(row.get("Costo Total"))
    return round(total, 2), method, f"{detail_key}; {len(rows)} filas en Hoja 1"


def fill_target_copy(indexes: dict[str, Any], timestamp: str) -> tuple[Path, Counter]:
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_path = OUTPUT_DIR / f"HOJA RESUMEN FINANCIERO OTS_FILLED_{timestamp}.xlsx"
    shutil.copy2(TARGET_FILE, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb[DESTINO_SHEET]
    ws.delete_rows(1, ws.max_row)
    ws.append(DEST_HEADERS)

    stats: Counter = Counter()
    for ot in sorted(indexes["ordenes_by_ot"]):
        orden_row = indexes["ordenes_by_ot"][ot]
        sp_match = match_sp(ot, orden_row, indexes)
        matriz_rows = sp_match.matriz_rows

        real_delivery = orden_row.get("FECHA REAL ENTREGA")
        request_date = orden_row.get("Marca temporal") or orden_row.get("FECHA DE SOLICITUD")
        execution_time = orden_row.get("TIEMPO ENTREGA (dias)") or days_between(request_date, real_delivery)

        labor_value, labor_method, labor_detail = labor_for_ot(ot, indexes)
        purchase_value, purchase_method, purchase_detail = purchase_value_for_ot(ot, sp_match.sp, indexes)

        reception_date = first_value(matriz_rows, "Fecha de Recepción de la SP  Nota: si no tiene fecha coloque la de la SP *")
        purchase_order_date = first_value(matriz_rows, "FECHA ORDEN DE COMPRA")
        approval_date = first_value(matriz_rows, "FECHA APROBACION")
        purchase_time = days_between(reception_date, purchase_order_date)
        approval_time = days_between(purchase_order_date, approval_date)

        ws.append(
            [
                ot,
                build_ot_status_value(orden_row),
                execution_time,
                labor_value,
                sp_match.sp,
                reception_date,
                purchase_time,
                first_value(matriz_rows, "ORDENES DE COMPRA"),
                approval_time,
                first_value(matriz_rows, "Estado Actual de la SP*"),
                purchase_value,
                sp_match.method,
                sp_match.detail,
                labor_method,
                labor_detail,
                purchase_method,
                purchase_detail,
            ]
        )

        stats[f"sp:{sp_match.method}"] += 1
        stats[f"labor:{labor_method}"] += 1
        stats[f"purchase:{purchase_method}"] += 1

    wb.save(output_path)
    return output_path, stats


def annotate_source_copy(source_file: Path, sheet_name: str, timestamp: str, indexes: dict[str, Any]) -> Path:
    output_path = OUTPUT_DIR / f"{source_file.stem}_ANNOTATED_{timestamp}.xlsx"
    shutil.copy2(source_file, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb[sheet_name]
    headers = [clean_header(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    source_max_col = len(headers)
    start_col = source_max_col + 1
    ws.cell(1, start_col).value = "CRUCE_RESUMEN"
    ws.cell(1, start_col + 1).value = "METODO_CRUCE_RESUMEN"
    ws.cell(1, start_col + 2).value = "DETALLE_CRUCE_RESUMEN"

    ordenes_by_ot = indexes["ordenes_by_ot"]
    matriz_by_sp = indexes["matriz_by_sp"]

    for row_number in range(2, ws.max_row + 1):
        row_values = {
            headers[col - 1]: ws.cell(row_number, col).value
            for col in range(1, source_max_col + 1)
            if headers[col - 1]
        }
        status = "NO_CRUZA"
        method = "sin_cruce"
        detail = ""

        if source_file == ORDENES_FILE:
            ot = normalize_ot(row_values.get("OT"), row_number) or normalize_ot(row_values.get("5"), row_number)
            if ot and ot in ordenes_by_ot:
                match = match_sp(ot, ordenes_by_ot[ot], indexes)
                status = "CRUZA"
                method = match.method
                detail = match.detail
        elif source_file == MATRIZ_FILE:
            ot = normalize_ot(row_values.get("OT"))
            sp = normalize_sp(row_values.get("NUMERO DE LA SP (solo el numero sin letras)*"))
            if ot and ot in ordenes_by_ot:
                status = "CRUZA"
                method = "directo_matriz_ot"
                detail = f"Matriz.OT={ot} existe en Ordenes"
            elif sp:
                matched_ots = []
                for candidate_ot, orden_row in ordenes_by_ot.items():
                    mentions = parse_sp_mentions(orden_row.get("COMENTARIOS"), orden_row.get("DESCRIPCIÓN GENERAL DEL FALLO O DE LA SOLICTUD"))
                    if sp in mentions:
                        matched_ots.append(candidate_ot)
                if matched_ots:
                    status = "CRUZA"
                    method = "truco_sp_en_texto"
                    detail = f"SP {sp} aparece en texto de OTs {matched_ots[:5]}"
                elif sp in matriz_by_sp:
                    status = "NO_CRUZA"
                    method = "sp_sin_ot_detectada"
                    detail = "SP existe en Matriz, pero no se detecto OT relacionada"
        elif source_file == ACTIVIDADES_FILE:
            ot = normalize_ot(row_values.get("OT")) or normalize_ot(row_values.get("ORDEN DE TRABAJO/REPORTE DE CAMPO"))
            if ot and ot in ordenes_by_ot:
                status = "CRUZA"
                method = "directo_actividades_ot"
                detail = f"Actividad.OT={ot} existe en Ordenes"
            elif ot:
                status = "NO_CRUZA"
                method = "ot_actividad_no_existe_en_ordenes"
                detail = f"Actividad.OT={ot} no existe en Ordenes"

        ws.cell(row_number, start_col).value = status
        ws.cell(row_number, start_col + 1).value = method
        ws.cell(row_number, start_col + 2).value = detail

    wb.save(output_path)
    return output_path


def write_report(timestamp: str, output_path: Path, annotated_paths: list[Path], stats: Counter, indexes: dict[str, Any]) -> Path:
    report_path = OUTPUT_DIR / f"RESUMEN_CRUCES_{timestamp}.md"
    total_ots = len(indexes["ordenes_by_ot"])
    lines = [
        "# Resumen de cruces",
        "",
        f"Archivo llenado: `{output_path.name}`",
        "",
        "## Archivos fuente anotados",
        *[f"- `{path.name}`" for path in annotated_paths],
        "",
        "## Estadisticas de SP",
    ]
    for key, count in sorted((k, v) for k, v in stats.items() if k.startswith("sp:")):
        lines.append(f"- `{key.replace('sp:', '')}`: {count}")
    lines.extend(["", "## Estadisticas de mano de obra"])
    for key, count in sorted((k, v) for k, v in stats.items() if k.startswith("labor:")):
        lines.append(f"- `{key.replace('labor:', '')}`: {count}")
    lines.extend(["", "## Estadisticas de valor de compra"])
    for key, count in sorted((k, v) for k, v in stats.items() if k.startswith("purchase:")):
        lines.append(f"- `{key.replace('purchase:', '')}`: {count}")

    lines.extend(
        [
            "",
            "## Totales",
            f"- OTs procesadas: {total_ots}",
            f"- OTs sin cruce de SP: {stats.get('sp:sin_cruce_sp', 0)}",
            f"- OTs sin actividades para mano de obra: {stats.get('labor:sin_actividades', 0)}",
            f"- OTs sin valor de compra en Hoja 1: {stats.get('purchase:sin_valor_compra', 0)}",
            "",
        ]
    )
    report_path.write_text("\n".join(lines), encoding="utf-8")
    return report_path


def main() -> None:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    indexes = build_indexes()
    output_path, stats = fill_target_copy(indexes, timestamp)
    annotated_paths = [
        annotate_source_copy(ORDENES_FILE, ORDENES_SHEET, timestamp, indexes),
        annotate_source_copy(MATRIZ_FILE, MATRIZ_SHEET, timestamp, indexes),
        annotate_source_copy(ACTIVIDADES_FILE, ACTIVIDADES_SHEET, timestamp, indexes),
    ]
    report_path = write_report(timestamp, output_path, annotated_paths, stats, indexes)

    print(f"Archivo llenado: {output_path}")
    print(f"Reporte: {report_path}")
    print("Archivos fuente anotados:")
    for path in annotated_paths:
        print(f"- {path}")
    print("\nEstadisticas:")
    for key, value in sorted(stats.items()):
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
